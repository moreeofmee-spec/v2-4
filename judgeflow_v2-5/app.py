import os, json, random, string
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from flask_babel import Babel
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'judgeflow-secret-2024')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:////home/cecii/judgeflow/judgeflow.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['BABEL_DEFAULT_LOCALE'] = 'hu'
app.config['BABEL_SUPPORTED_LOCALES'] = ['hu', 'en']

db = SQLAlchemy(app)
babel = Babel(app)

# ── Models ───────────────────────────────────────────────────────────────────

class Competition(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    public_id = db.Column(db.String(10), unique=True, nullable=True)   # 5-digit public ID
    organizer_code = db.Column(db.String(10), unique=True, nullable=True)  # 4num+2letter
    name = db.Column(db.String(200), nullable=False)
    password = db.Column(db.String(100), nullable=False)
    organizer_name = db.Column(db.String(200))
    organizer_email = db.Column(db.String(200))
    organizer_phone = db.Column(db.String(50))
    date = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    events = db.relationship('Event', backref='competition', lazy=True, cascade='all, delete-orphan')
    judges = db.relationship('Judge', backref='competition', lazy=True, cascade='all, delete-orphan')

class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    competition_id = db.Column(db.Integer, db.ForeignKey('competition.id'), nullable=False)
    riders = db.relationship('Rider', backref='event', lazy=True, cascade='all, delete-orphan')
    categories = db.relationship('Category', backref='event', lazy=True, cascade='all, delete-orphan')
    licences = db.relationship('Licence', backref='event', lazy=True, cascade='all, delete-orphan')

class Judge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    qualification = db.Column(db.String(100))
    judge_code = db.Column(db.String(10))  # 4-digit code set by organizer
    competition_id = db.Column(db.Integer, db.ForeignKey('competition.id'), nullable=False)
    scores = db.relationship('Score', backref='judge', lazy=True, cascade='all, delete-orphan')


class EventTaskMultiplier(db.Model):
    """Stores which task numbers have a multiplier for a given event (custom events only)"""
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    task_number = db.Column(db.Integer, nullable=False)  # 1-30
    multiplier = db.Column(db.Float, default=2.0)

class EventJudge(db.Model):
    """Links a judge to a specific event with a position"""
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    judge_id = db.Column(db.Integer, db.ForeignKey('judge.id'), nullable=False)
    position = db.Column(db.String(10))  # E, H, C, M, B - per event

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)

class Licence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)

class Rider(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    horse = db.Column(db.String(200))
    licence = db.Column(db.String(100))
    category = db.Column(db.String(100))
    start_number = db.Column(db.Integer)
    email = db.Column(db.String(200))
    phone = db.Column(db.String(50))
    rider_code = db.Column(db.String(10))  # birth year+day default, changeable
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    scores = db.relationship('Score', backref='rider', lazy=True, cascade='all, delete-orphan')

class Score(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rider_id = db.Column(db.Integer, db.ForeignKey('rider.id'), nullable=False)
    judge_id = db.Column(db.Integer, db.ForeignKey('judge.id'), nullable=False)
    tasks = db.Column(db.Text)
    error_points = db.Column(db.Float, default=0.0)
    multiplier = db.Column(db.Float, default=1.0)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)

# ── Babel ─────────────────────────────────────────────────────────────────────

def get_locale():
    return session.get('lang', 'hu')

babel.init_app(app, locale_selector=get_locale)

# ── Constants ─────────────────────────────────────────────────────────────────

DEFAULT_CATEGORIES = ['Gyerek / Child', 'Fiatal ló / Young Horse', 'U25', 'Felnőtt / Adult']
DEFAULT_LICENCES = ['A', 'B', 'C', 'D']
JUDGE_POSITIONS = ['E', 'H', 'C', 'M', 'B']

# ── FEI Dressage Tests 2026 ───────────────────────────────────────────────────
# Source: inside.fei.org/fei/your-role/organisers/dressage/tests
# Updated: January 2026

FEI_TESTS = {
    "Seniors": [
        "Prix St-Georges",
        "Intermediate I",
        "Intermediate A",
        "Intermediate B",
        "Intermediate II",
        "Grand Prix",
        "Grand Prix Special",
        "Grand Prix Freestyle",
        "Short Grand Prix (CDI-Ws only)",
    ],
    "U25": [
        "U25 Grand Prix",
        "U25 Grand Prix Freestyle",
    ],
    "Young Riders": [
        "Young Riders Team Test",
        "Young Riders Individual Test",
        "Young Riders Freestyle",
        "Young Riders Preliminary Test",
    ],
    "Juniors": [
        "Junior Team Test",
        "Junior Individual Test",
        "Junior Freestyle",
        "Junior Preliminary Test",
    ],
    "Pony Riders": [
        "Pony Rider Team Test",
        "Pony Rider Individual Test",
        "Pony Rider Freestyle",
        "Pony Rider Preliminary Test",
    ],
    "Children": [
        "Children Team Test",
        "Children Individual Test",
        "Children Preliminary Test",
    ],
    "Young Horses": [
        "5-Year-Old Preliminary Test",
        "5-Year-Old Final Test",
        "6-Year-Old Preliminary Test",
        "6-Year-Old Final Test",
        "7-Year-Old Preliminary Test",
        "7-Year-Old Final Test",
    ],
}

# Flat list for lookup + "Custom" option
FEI_TESTS_FLAT = []
for cat, tests in FEI_TESTS.items():
    for t in tests:
        FEI_TESTS_FLAT.append(f"{cat} – {t}")



def hu():
    return session.get('lang', 'hu') == 'hu'

def gen_public_id():
    while True:
        pid = str(random.randint(10000, 99999))
        if not Competition.query.filter_by(public_id=pid).first():
            return pid

def gen_organizer_code():
    while True:
        nums = ''.join([str(random.randint(0,9)) for _ in range(4)])
        lets = ''.join(random.choices(string.ascii_uppercase, k=2))
        code = nums + lets
        if not Competition.query.filter_by(organizer_code=code).first():
            return code

def send_admin_notification(comp_name, organizer_name, organizer_email, date):
    admin_email = os.environ.get('ADMIN_EMAIL')
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    if not all([admin_email, smtp_user, smtp_pass]):
        return
    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = admin_email
        msg['Subject'] = f'JudgeFlow – Uj verseny: {comp_name}'
        body = f"Verseny: {comp_name}\nRendezo: {organizer_name}\nEmail: {organizer_email}\nDatum: {date}"
        msg.attach(MIMEText(body, 'plain'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, admin_email, msg.as_string())
    except Exception as e:
        print(f"[JudgeFlow] Email error: {e}")

def send_code_email(to_email, subject, body):
    """Send code reminder email via Gmail SMTP"""
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    if not all([smtp_user, smtp_pass, to_email]):
        return False
    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"[JudgeFlow] Email send error: {e}")
        return False


def calc_score(score_obj, task_multipliers=None):
    """Calculate score. task_multipliers = {task_index: multiplier} (0-indexed)"""
    if not score_obj or not score_obj.tasks:
        return 0.0
    tasks = json.loads(score_obj.tasks)
    total = 0.0
    max_possible = 0.0
    for i, t in enumerate(tasks):
        try:
            val = float(t) if t not in [None, ''] else 0.0
        except:
            val = 0.0
        task_mult = 1.0
        if task_multipliers and (i+1) in task_multipliers:
            task_mult = task_multipliers[i+1]
        total += val * task_mult
        max_possible += 10.0 * task_mult
    # Apply error and overall multiplier
    result = (total - score_obj.error_points) * score_obj.multiplier
    max_score = max_possible * score_obj.multiplier
    return max(0.0, min(result, max_score))

def get_max_score(event_id):
    """Get max possible score for an event considering multipliers"""
    mults = EventTaskMultiplier.query.filter_by(event_id=event_id).all()
    mult_map = {m.task_number: m.multiplier for m in mults}
    total = sum(10.0 * mult_map.get(i, 1.0) for i in range(1, 31))
    return total  # default 300 (30 tasks x 10pts)

def calc_percent(val, max_pts=300):
    if max_pts <= 0: return 0.0
    return round(min((val / max_pts) * 100, 100.0), 2)  # never exceed 100%

def get_results(event_id, filter_category=None, filter_licence=None, filter_judge_id=None):
    riders = Rider.query.filter_by(event_id=event_id).all()
    event = Event.query.get(event_id)
    if not event:
        return []
    # Get judges assigned to this event with their positions
    event_judges = EventJudge.query.filter_by(event_id=event_id).all()
    judge_map = {ej.judge_id: ej.position for ej in event_judges}
    judges = [Judge.query.get(ej.judge_id) for ej in event_judges]
    judges = [j for j in judges if j]
    results = []
    # Get task multipliers for this event
    mults = EventTaskMultiplier.query.filter_by(event_id=event_id).all()
    task_multipliers = {m.task_number: m.multiplier for m in mults}
    max_pts = get_max_score(event_id)

    for rider in riders:
        if filter_category and rider.category != filter_category:
            continue
        if filter_licence and rider.licence != filter_licence:
            continue
        judge_scores = []
        for judge in judges:
            if filter_judge_id and judge.id != int(filter_judge_id):
                continue
            s = Score.query.filter_by(rider_id=rider.id, judge_id=judge.id).first()
            if s:
                raw = calc_score(s, task_multipliers)
                judge_scores.append({'judge': judge.name, 'judge_id': judge.id,
                                     'position': judge_map.get(judge.id, ''), 'raw': raw,
                                     'percent': calc_percent(raw, max_pts)})
        if judge_scores:
            avg = sum(j['raw'] for j in judge_scores) / len(judge_scores)
            avg = min(avg, max_pts)  # max_pts default 300
            results.append({'rider': rider, 'judge_scores': judge_scores,
                            'avg_raw': round(avg, 2), 'avg_percent': calc_percent(avg, max_pts)})
    results.sort(key=lambda x: x['avg_percent'], reverse=True)
    for i, r in enumerate(results):
        r['rank'] = i + 1
    return results

# ── Lang ──────────────────────────────────────────────────────────────────────

@app.route('/lang/<lang>')
def set_lang(lang):
    if lang in ['hu', 'en']:
        session['lang'] = lang
    return redirect(request.referrer or url_for('index'))

# ── Index ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

# ── Organizer ─────────────────────────────────────────────────────────────────

@app.route('/organizer')
def organizer_home():
    return render_template('organizer_home.html')

@app.route('/organizer/create', methods=['GET', 'POST'])
def create_competition():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        organizer_name = request.form.get('organizer_name', '').strip()
        organizer_email = request.form.get('organizer_email', '').strip()
        organizer_phone = request.form.get('organizer_phone', '').strip()
        date = request.form.get('date', '').strip()
        errors = []
        if not name: errors.append('Verseny neve kötelező!' if hu() else 'Name required!')
        if not organizer_name: errors.append('Rendező neve kötelező!' if hu() else 'Organizer name required!')
        if not organizer_email: errors.append('Email kötelező!' if hu() else 'Email required!')
        if not organizer_phone: errors.append('Telefon kötelező!' if hu() else 'Phone required!')
        if errors:
            for e in errors: flash(e, 'error')
            return render_template('create_competition.html')
        pub_id = gen_public_id()
        org_code = gen_organizer_code()
        # No separate password needed - organizer_code IS the auth
        comp = Competition(name=name, password=org_code, public_id=pub_id, organizer_code=org_code,
                           organizer_name=organizer_name, organizer_email=organizer_email,
                           organizer_phone=organizer_phone, date=date)
        db.session.add(comp)
        db.session.commit()
        send_admin_notification(name, organizer_name, organizer_email, date)
        session[f'organizer_{comp.id}'] = True
        # Send organizer email notification with codes
        send_code_email(
            organizer_email,
            f'JudgeFlow – Verseny létrehozva: {name}',
            f"""JudgeFlow – Verseny sikeresen létrehozva!

Verseny neve: {name}
Verseny azonosító (indulóknak): {pub_id}
Rendező kód (neked): {org_code}
Dátum: {date}

Ezt a rendező kódot tárold el biztonságos helyen - ezzel tudsz visszalépni szerkeszteni.
A verseny azonosítót add meg az indulóknak az eredmények megtekintéséhez.

JudgeFlow csapat"""
        )
        flash((f'Verseny létrehozva! Azonosító: {pub_id} | Rendező kód: {org_code} (emailben is elküldve)' if hu()
               else f'Competition created! ID: {pub_id} | Organizer code: {org_code} (sent by email)'), 'success')
        return redirect(url_for('manage_competition', comp_id=comp.id))
    return render_template('create_competition.html')

@app.route('/organizer/login', methods=['GET', 'POST'])
def organizer_login():
    if request.method == 'POST':
        org_code = request.form.get('organizer_code', '').strip().upper()
        comp = Competition.query.filter_by(organizer_code=org_code).first()
        if comp:
            session[f'organizer_{comp.id}'] = True
            return redirect(url_for('manage_competition', comp_id=comp.id))
        flash('Hibás rendező kód!' if hu() else 'Wrong organizer code!', 'error')
    return render_template('organizer_login.html')

@app.route('/competition/<int:comp_id>/manage')
def manage_competition(comp_id):
    if not session.get(f'organizer_{comp_id}'):
        return redirect(url_for('organizer_login'))
    comp = Competition.query.get_or_404(comp_id)
    events = Event.query.filter_by(competition_id=comp_id).all()
    judges = Judge.query.filter_by(competition_id=comp_id).all()
    # Count judges per event
    event_judge_counts = {}
    for event in events:
        event_judge_counts[event.id] = EventJudge.query.filter_by(event_id=event.id).count()
    return render_template('manage.html', comp=comp, events=events, judges=judges,
                           event_judge_counts=event_judge_counts,
                           judge_positions=JUDGE_POSITIONS,
                           fei_tests=FEI_TESTS, fei_tests_flat=FEI_TESTS_FLAT)

@app.route('/competition/<int:comp_id>/add-event', methods=['POST'])
def add_event(comp_id):
    if not session.get(f'organizer_{comp_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    fei_type = data.get('fei_test_type', '')
    is_custom = (fei_type == 'custom' or not fei_type)
    display_name = data['name'] if is_custom else (fei_type if not data.get('name') else data['name'])
    event = Event(name=display_name, fei_test_type=fei_type if not is_custom else None,
                  is_custom=is_custom, competition_id=comp_id)
    db.session.add(event)
    db.session.flush()
    for cat in DEFAULT_CATEGORIES:
        db.session.add(Category(name=cat, event_id=event.id))
    for lic in DEFAULT_LICENCES:
        db.session.add(Licence(name=lic, event_id=event.id))
    db.session.commit()
    return jsonify({'id': event.id, 'name': event.name, 'is_custom': is_custom})

@app.route('/event/<int:event_id>/manage')
def manage_event(event_id):
    event = Event.query.get_or_404(event_id)
    comp_id = event.competition_id
    if not session.get(f'organizer_{comp_id}'):
        return redirect(url_for('organizer_login'))
    all_judges = Judge.query.filter_by(competition_id=comp_id).all()
    event_judges = EventJudge.query.filter_by(event_id=event_id).all()
    event_judge_ids = {ej.judge_id: ej for ej in event_judges}
    riders = Rider.query.filter_by(event_id=event_id).order_by(Rider.start_number).all()
    categories = Category.query.filter_by(event_id=event_id).all()
    licences = Licence.query.filter_by(event_id=event_id).all()
    multipliers = {}
    if event.is_custom:
        mults = EventTaskMultiplier.query.filter_by(event_id=event_id).all()
        multipliers = {m.task_number: m.multiplier for m in mults}
    return render_template('manage_event.html', event=event, comp=event.competition,
                           all_judges=all_judges, event_judges=event_judges,
                           event_judge_ids=event_judge_ids,
                           riders=riders, categories=categories,
                           licences=licences, judge_positions=JUDGE_POSITIONS,
                           multipliers=multipliers)

# ── Judge/Rider CRUD ──────────────────────────────────────────────────────────

@app.route('/competition/<int:comp_id>/add-judge', methods=['POST'])
def add_judge(comp_id):
    if not session.get(f'organizer_{comp_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'name required'}), 400
    judge = Judge(name=data['name'].strip(),
                  qualification=data.get('qualification', ''),
                  judge_code=data.get('judge_code', ''), competition_id=comp_id)
    db.session.add(judge)
    db.session.commit()
    return jsonify({'id': judge.id, 'name': judge.name,
                    'qualification': judge.qualification, 'judge_code': judge.judge_code})

@app.route('/competition/<int:comp_id>/edit-judge/<int:judge_id>', methods=['POST'])
def edit_judge(comp_id, judge_id):
    if not session.get(f'organizer_{comp_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    judge = Judge.query.get_or_404(judge_id)
    data = request.get_json()
    judge.name = data.get('name', judge.name)
    judge.qualification = data.get('qualification', judge.qualification)
    judge.judge_code = data.get('judge_code', judge.judge_code)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/competition/<int:comp_id>/delete-judge/<int:judge_id>', methods=['POST'])
def delete_judge(comp_id, judge_id):
    if not session.get(f'organizer_{comp_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    db.session.delete(Judge.query.get_or_404(judge_id))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/event/<int:event_id>/add-rider', methods=['POST'])
def add_rider(event_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    existing = Rider.query.filter_by(event_id=event_id).order_by(Rider.start_number.desc()).first()
    next_start = (existing.start_number or 0) + 1 if existing else 1
    # default rider_code = birth_year + birth_day (set by organizer, e.g. 200508)
    rider_code = data.get('rider_code', '')
    rider = Rider(name=data['name'], horse=data.get('horse', ''), licence=data.get('licence', ''),
                  category=data.get('category', ''), start_number=next_start,
                  email=data.get('email', ''), phone=data.get('phone', ''),
                  rider_code=rider_code, event_id=event_id)
    db.session.add(rider)
    db.session.commit()
    return jsonify({'id': rider.id, 'name': rider.name, 'start_number': rider.start_number})

@app.route('/event/<int:event_id>/edit-rider/<int:rider_id>', methods=['POST'])
def edit_rider(event_id, rider_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    rider = Rider.query.get_or_404(rider_id)
    data = request.get_json()
    rider.name = data.get('name', rider.name)
    rider.horse = data.get('horse', rider.horse)
    rider.licence = data.get('licence', rider.licence)
    rider.category = data.get('category', rider.category)
    rider.start_number = data.get('start_number', rider.start_number)
    rider.email = data.get('email', rider.email)
    rider.phone = data.get('phone', rider.phone)
    rider.rider_code = data.get('rider_code', rider.rider_code)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/event/<int:event_id>/delete-rider/<int:rider_id>', methods=['POST'])
def delete_rider(event_id, rider_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    db.session.delete(Rider.query.get_or_404(rider_id))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/event/<int:event_id>/add-event-judge', methods=['POST'])
def add_event_judge(event_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    # Check not already added
    existing = EventJudge.query.filter_by(event_id=event_id, judge_id=data['judge_id']).first()
    if existing:
        existing.position = data.get('position', existing.position)
        db.session.commit()
        return jsonify({'ok': True, 'updated': True})
    ej = EventJudge(event_id=event_id, judge_id=int(data['judge_id']),
                    position=data.get('position', ''))
    db.session.add(ej)
    db.session.commit()
    return jsonify({'ok': True, 'id': ej.id})

@app.route('/event/<int:event_id>/remove-event-judge/<int:judge_id>', methods=['POST'])
def remove_event_judge(event_id, judge_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    ej = EventJudge.query.filter_by(event_id=event_id, judge_id=judge_id).first()
    if ej:
        db.session.delete(ej)
        db.session.commit()
    return jsonify({'ok': True})

@app.route('/event/<int:event_id>/add-category', methods=['POST'])
def add_category(event_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    cat = Category(name=data['name'], event_id=event_id)
    db.session.add(cat)
    db.session.commit()
    return jsonify({'id': cat.id, 'name': cat.name})

@app.route('/event/<int:event_id>/add-licence', methods=['POST'])
def add_licence(event_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    lic = Licence(name=data['name'], event_id=event_id)
    db.session.add(lic)
    db.session.commit()
    return jsonify({'id': lic.id, 'name': lic.name})

# ── Judge login (separate from organizer) ─────────────────────────────────────

@app.route('/judge')
def judge_home():
    return render_template('judge_home.html')

@app.route('/judge/login', methods=['GET', 'POST'])
def judge_login_home():
    """Judge enters competition public_id first, then selects their name"""
    if request.method == 'POST':
        public_id = request.form.get('public_id', '').strip()
        comp = Competition.query.filter_by(public_id=public_id).first()
        if not comp:
            flash('Nem található verseny!' if hu() else 'Competition not found!', 'error')
            return render_template('judge_home.html')
        events = Event.query.filter_by(competition_id=comp.id).all()
        return render_template('judge_select_event.html', comp=comp, events=events)
    return render_template('judge_home.html')

@app.route('/judge/select-event/<int:comp_id>')
def judge_select_event(comp_id):
    comp = Competition.query.get_or_404(comp_id)
    judge_id = session.get(f'judge_comp_{comp_id}')
    if not judge_id:
        return redirect(url_for('judge_login_comp', comp_id=comp_id))
    judge = Judge.query.get_or_404(judge_id)
    events = Event.query.filter_by(competition_id=comp_id).all()
    # Sort: events where this judge has scores first
    scored_event_ids = set(s.judge_id and Score.query.filter_by(judge_id=judge_id).first() and
                           Rider.query.get(Score.query.filter_by(judge_id=judge_id).first().rider_id).event_id
                           for s in Score.query.filter_by(judge_id=judge_id).all()
                           if s)
    return render_template('judge_events.html', comp=comp, events=events, judge=judge)

@app.route('/judge/login/<int:comp_id>', methods=['GET', 'POST'])
def judge_login_comp(comp_id):
    comp = Competition.query.get_or_404(comp_id)
    judges = Judge.query.filter_by(competition_id=comp_id).all()
    if request.method == 'POST':
        judge_id = request.form.get('judge_id', '')
        judge_code = request.form.get('judge_code', '').strip()
        if judge_id:
            judge = Judge.query.get(judge_id)
            if judge and judge.judge_code and judge.judge_code == judge_code:
                session[f'judge_comp_{comp_id}'] = int(judge_id)
                return redirect(url_for('judge_select_event', comp_id=comp_id))
            elif judge and not judge.judge_code:
                session[f'judge_comp_{comp_id}'] = int(judge_id)
                return redirect(url_for('judge_select_event', comp_id=comp_id))
        flash('Hibás kód!' if hu() else 'Wrong code!', 'error')
    return render_template('judge_login_comp.html', comp=comp, judges=judges)

@app.route('/event/<int:event_id>/judge-login', methods=['GET', 'POST'])
def judge_login(event_id):
    """Legacy redirect - now handled at competition level"""
    event = Event.query.get_or_404(event_id)
    return redirect(url_for('judge_login_comp', comp_id=event.competition_id))

@app.route('/event/<int:event_id>/judge')
def judge_panel(event_id):
    event = Event.query.get_or_404(event_id)
    comp_id = event.competition_id
    judge_id = session.get(f'judge_comp_{comp_id}')
    if not judge_id:
        return redirect(url_for('judge_login_comp', comp_id=comp_id))
    judge = Judge.query.get_or_404(judge_id)
    riders = Rider.query.filter_by(event_id=event_id).order_by(Rider.start_number).all()
    scores = {s.rider_id: s for s in Score.query.filter_by(judge_id=judge_id).all()}
    return render_template('judge_panel.html', event=event, comp=event.competition,
                           judge=judge, riders=riders, scores=scores)

@app.route('/event/<int:event_id>/score/<int:rider_id>', methods=['GET', 'POST'])
def score_rider(event_id, rider_id):
    event = Event.query.get_or_404(event_id)
    judge_id = session.get(f'judge_comp_{event.competition_id}')
    if not judge_id:
        return redirect(url_for('judge_login_comp', comp_id=event.competition_id))
    judge = Judge.query.get_or_404(judge_id)
    rider = Rider.query.get_or_404(rider_id)
    score = Score.query.filter_by(rider_id=rider_id, judge_id=judge_id).first()
    if request.method == 'POST':
        tasks = [request.form.get(f'task_{i}', '') for i in range(1, 31)]
        # Validate: empty task = error
        empty_tasks = [i+1 for i, t in enumerate(tasks) if t.strip() == '']
        if empty_tasks:
            flash(f'Hiányzó pontszám a következő feladatoknál: {", ".join(map(str, empty_tasks[:5]))}{"..." if len(empty_tasks)>5 else ""}' if hu() else f'Missing scores for tasks: {", ".join(map(str, empty_tasks[:5]))}{"..." if len(empty_tasks)>5 else ""}', 'error')
            existing_tasks = tasks
            return render_template('score_form.html', event=event, comp=event.competition,
                                   judge=judge, rider=rider, score=score, existing_tasks=existing_tasks)
        # clean tasks, cap at 10
        cleaned = []
        for t in tasks:
            try:
                v = float(t)
                cleaned.append(str(min(max(v, 0), 10)))
            except:
                cleaned.append('0')
        error_points = float(request.form.get('error_points', 0) or 0)
        multiplier = float(request.form.get('multiplier', 1.0) or 1.0)
        if score:
            score.tasks = json.dumps(cleaned)
            score.error_points = error_points
            score.multiplier = multiplier
            score.submitted_at = datetime.utcnow()
        else:
            score = Score(rider_id=rider_id, judge_id=judge_id,
                          tasks=json.dumps(cleaned), error_points=error_points, multiplier=multiplier)
            db.session.add(score)
        db.session.commit()
        flash('Pontszámok elmentve!' if hu() else 'Scores saved!', 'success')
        return redirect(url_for('judge_panel', event_id=event_id))
    existing_tasks = json.loads(score.tasks) if score and score.tasks else [''] * 30
    return render_template('score_form.html', event=event, comp=event.competition,
                           judge=judge, rider=rider, score=score, existing_tasks=existing_tasks)

# ── Rider (versenyző) flow ────────────────────────────────────────────────────

@app.route('/rider')
def rider_home():
    return render_template('rider_home.html')

@app.route('/rider/enter', methods=['POST'])
def rider_enter():
    comp_id_input = request.form.get('comp_id', '').strip()
    comp = Competition.query.filter_by(public_id=comp_id_input).first()
    if not comp:
        flash('Nem található verseny ezzel az azonosítóval.' if hu() else 'No competition found.', 'error')
        return redirect(url_for('rider_home'))
    events = Event.query.filter_by(competition_id=comp.id).all()
    return render_template('rider_events.html', comp=comp, events=events)

@app.route('/rider/event/<int:event_id>/login', methods=['GET', 'POST'])
def rider_event_login(event_id):
    """Rider selects name, confirms with start_number if duplicates, then enters code"""
    event = Event.query.get_or_404(event_id)
    riders = Rider.query.filter_by(event_id=event_id).order_by(Rider.name).all()
    if request.method == 'POST':
        rider_id = request.form.get('rider_id', '')
        rider_code = request.form.get('rider_code', '').strip()
        start_number = request.form.get('start_number', '').strip()
        rider = Rider.query.get(rider_id)
        if not rider:
            flash('Nincs ilyen versenyző!' if hu() else 'Rider not found!', 'error')
            return render_template('rider_event_login.html', event=event, comp=event.competition, riders=riders)
        # check for duplicate names
        same_name = Rider.query.filter_by(event_id=event_id, name=rider.name).all()
        if len(same_name) > 1 and str(rider.start_number) != start_number:
            flash('Kérjük add meg a rajtszámodat is!' if hu() else 'Please confirm your start number!', 'error')
            return render_template('rider_event_login.html', event=event, comp=event.competition,
                                   riders=riders, show_start=rider_id)
        if rider.rider_code and rider.rider_code != rider_code:
            flash('Hibás kód!' if hu() else 'Wrong code!', 'error')
            return render_template('rider_event_login.html', event=event, comp=event.competition, riders=riders)
        session[f'rider_{event_id}'] = rider.id
        return redirect(url_for('rider_results', event_id=event_id))
    return render_template('rider_event_login.html', event=event, comp=event.competition, riders=riders)

@app.route('/rider/event/<int:event_id>/results')
def rider_results(event_id):
    rider_id = session.get(f'rider_{event_id}')
    event = Event.query.get_or_404(event_id)
    res = get_results(event_id)
    categories = Category.query.filter_by(event_id=event_id).all()
    licences = Licence.query.filter_by(event_id=event_id).all()
    judges = Judge.query.filter_by(event_id=event_id).all()
    filter_cat = request.args.get('category', '')
    filter_lic = request.args.get('licence', '')
    if filter_cat or filter_lic:
        res = get_results(event_id, filter_cat or None, filter_lic or None)
    return render_template('rider_results.html', event=event, comp=event.competition,
                           results=res, my_rider_id=rider_id,
                           categories=categories, licences=licences, judges=judges,
                           filter_cat=filter_cat, filter_lic=filter_lic)

@app.route('/rider/change-code', methods=['POST'])
def rider_change_code():
    rider_id = request.form.get('rider_id')
    event_id = request.form.get('event_id')
    new_code = request.form.get('new_code', '').strip()
    rider = Rider.query.get(rider_id)
    if rider and len(new_code) == 6 and new_code.isdigit():
        rider.rider_code = new_code
        db.session.commit()
        flash('Kód megváltoztatva!' if hu() else 'Code updated!', 'success')
    else:
        flash('6 jegyű számot adj meg!' if hu() else 'Enter a 6-digit number!', 'error')
    return redirect(url_for('rider_results', event_id=event_id))

# ── Results (public) ──────────────────────────────────────────────────────────

@app.route('/event/<int:event_id>/results')
def results(event_id):
    event = Event.query.get_or_404(event_id)
    filter_cat = request.args.get('category', '')
    filter_lic = request.args.get('licence', '')
    filter_jid = request.args.get('judge_id', '')
    res = get_results(event_id, filter_cat or None, filter_lic or None, filter_jid or None)
    judges = Judge.query.filter_by(event_id=event_id).all()
    categories = Category.query.filter_by(event_id=event_id).all()
    licences = Licence.query.filter_by(event_id=event_id).all()
    return render_template('results.html', event=event, comp=event.competition, results=res,
                           judges=judges, categories=categories, licences=licences,
                           filter_cat=filter_cat, filter_lic=filter_lic, filter_jid=filter_jid,
                           my_rider_id=None)

# ── Export ────────────────────────────────────────────────────────────────────

@app.route('/event/<int:event_id>/export/excel')
def export_excel(event_id):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    event = Event.query.get_or_404(event_id)
    filter_cat = request.args.get('category', '')
    filter_lic = request.args.get('licence', '')
    filter_jid = request.args.get('judge_id', '')
    res = get_results(event_id, filter_cat or None, filter_lic or None, filter_jid or None)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'
    hfill = PatternFill('solid', fgColor='71C400')
    hfont = Font(bold=True, color='FFFFFF')
    headers = ['Rank','Start#','Rider','Horse','Category','Licence','Avg %'] + ['Judge','Pos','%']
    widths =  [8,      8,      30,     25,     18,        12,       10,       25,     6,    10]
    ws.append(['Rank','Start#','Rider','Horse','Category','Licence','Avg %','Judge','Pos','Judge %'])
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
    for col, w in zip(ws[1], [8,8,30,25,18,12,10,25,6,10]):
        ws.column_dimensions[cell.column_letter].width = w
    for r in res:
        first = True
        for js in r['judge_scores']:
            if first:
                ws.append([r['rank'], r['rider'].start_number, r['rider'].name, r['rider'].horse,
                            r['rider'].category, r['rider'].licence, r['avg_percent'],
                            js['judge'], js['position'], js['percent']])
                first = False
            else:
                ws.append(['','','','','','','', js['judge'], js['position'], js['percent']])
        if not r['judge_scores']:
            ws.append([r['rank'], r['rider'].start_number, r['rider'].name, r['rider'].horse,
                        r['rider'].category, r['rider'].licence, r['avg_percent'],'','',''])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f'judgeflow_{event.name}.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/event/<int:event_id>/export/pdf')
def export_pdf(event_id):
    import io
    from fpdf import FPDF
    event = Event.query.get_or_404(event_id)
    filter_cat = request.args.get('category', '')
    filter_lic = request.args.get('licence', '')
    filter_jid = request.args.get('judge_id', '')
    res = get_results(event_id, filter_cat or None, filter_lic or None, filter_jid or None)
    pdf = FPDF(); pdf.add_page()
    pdf.set_font('Helvetica','B',14)
    pdf.set_fill_color(113,196,0); pdf.set_text_color(255,255,255)
    pdf.cell(0,10,f'JudgeFlow - {event.competition.name} / {event.name}',ln=True,fill=True,align='C')
    pdf.set_text_color(0,0,0); pdf.set_font('Helvetica','',8); pdf.ln(3)
    cols = ['#','Rider','Horse','Cat','Lic','Avg%','Judge','Pos','J%']
    widths = [8,42,35,22,12,12,30,8,12]
    pdf.set_font('Helvetica','B',8); pdf.set_fill_color(113,196,0); pdf.set_text_color(255,255,255)
    for col,w in zip(cols,widths): pdf.cell(w,7,col,border=1,fill=True)
    pdf.ln(); pdf.set_font('Helvetica','',7); pdf.set_text_color(0,0,0)
    for i,r in enumerate(res):
        fill = i%2==0
        if fill: pdf.set_fill_color(240,248,230)
        first = True
        for js in r['judge_scores']:
            if first:
                row=[str(r['rank']),r['rider'].name,r['rider'].horse,r['rider'].category,
                     r['rider'].licence,str(r['avg_percent']),js['judge'],js['position'],str(js['percent'])]
                first=False
            else:
                row=['','','','','','',js['judge'],js['position'],str(js['percent'])]
            for val,w in zip(row,widths): pdf.cell(w,6,str(val)[:20],border=1,fill=fill)
            pdf.ln()
        if not r['judge_scores']:
            row=[str(r['rank']),r['rider'].name,r['rider'].horse,r['rider'].category,
                 r['rider'].licence,str(r['avg_percent']),'','','']
            for val,w in zip(row,widths): pdf.cell(w,6,str(val)[:20],border=1,fill=fill)
            pdf.ln()
    buf=io.BytesIO(pdf.output()); buf.seek(0)
    return send_file(buf,download_name=f'judgeflow_{event.name}.pdf',as_attachment=True,mimetype='application/pdf')



# ── Logout routes ─────────────────────────────────────────────────────────────

@app.route('/logout/organizer/<int:comp_id>')
def logout_organizer(comp_id):
    session.pop(f'organizer_{comp_id}', None)
    return redirect(url_for('index'))

@app.route('/logout/judge/<int:comp_id>')
def logout_judge(comp_id):
    session.pop(f'judge_comp_{comp_id}', None)
    return redirect(url_for('index'))

@app.route('/logout/rider/<int:event_id>')
def logout_rider(event_id):
    session.pop(f'rider_{event_id}', None)
    return redirect(url_for('index'))


@app.route('/event/<int:event_id>/set-multipliers', methods=['POST'])
def set_multipliers(event_id):
    event = Event.query.get_or_404(event_id)
    if not session.get(f'organizer_{event.competition_id}'):
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json()
    # data = {'multipliers': {task_num: multiplier_value, ...}}
    # Clear existing
    EventTaskMultiplier.query.filter_by(event_id=event_id).delete()
    for task_str, mult in data.get('multipliers', {}).items():
        try:
            task_num = int(task_str)
            mult_val = float(mult)
            if mult_val != 1.0 and 1 <= task_num <= 30:
                db.session.add(EventTaskMultiplier(event_id=event_id, task_number=task_num, multiplier=mult_val))
        except:
            pass
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/event/<int:event_id>/get-multipliers')
def get_multipliers(event_id):
    mults = EventTaskMultiplier.query.filter_by(event_id=event_id).all()
    return jsonify({str(m.task_number): m.multiplier for m in mults})

# ── Forgot code routes ────────────────────────────────────────────────────────

@app.route('/organizer/forgot', methods=['GET', 'POST'])
def organizer_forgot():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        # Find competitions by organizer email
        comps = Competition.query.filter(
            db.func.lower(Competition.organizer_email) == email
        ).all()
        if comps:
            body_lines = ["JudgeFlow – Rendező kód emlékeztető\n"]
            for c in comps:
                body_lines.append(f"Verseny: {c.name}")
                body_lines.append(f"Rendező kód: {c.organizer_code}")
                body_lines.append(f"Verseny ID: {c.public_id}\n")
            body = "\n".join(body_lines)
            sent = send_code_email(
                email,
                "JudgeFlow – Rendező kód emlékeztető",
                body
            )
            if sent:
                flash("Email elküldve!" if hu() else "Email sent!", "success")
            else:
                flash("Nem sikerült elküldeni. Ellenőrizd az email beállításokat." if hu() else "Could not send email.", "error")
        else:
            flash("Nem található verseny ezzel az email címmel." if hu() else "No competition found with this email.", "error")
        return redirect(url_for('organizer_login'))
    return render_template('forgot_organizer.html')


@app.route('/rider/forgot', methods=['GET', 'POST'])
def rider_forgot():
    """Rider forgot code - send to registered email or phone"""
    event_id = request.args.get('event_id', '')
    if request.method == 'POST':
        event_id = request.form.get('event_id', '')
        contact = request.form.get('contact', '').strip().lower()
        event = Event.query.get(event_id) if event_id else None
        if not event:
            flash("Nem található versenyszám." if hu() else "Event not found.", "error")
            return redirect(url_for('rider_home'))
        # Find rider by email or phone in this event
        riders = Rider.query.filter_by(event_id=event_id).all()
        found = None
        for r in riders:
            if (r.email and r.email.lower() == contact) or (r.phone and r.phone == contact):
                found = r
                break
        if found:
            code = found.rider_code or "Nincs kód beállítva"
            body = f"""JudgeFlow – Versenyző kód emlékeztető

Versenyző: {found.name}
Verseny: {event.competition.name} / {event.name}
Rajtszám: {found.start_number}
Kódod: {code}

Ha meg szeretnéd változtatni a kódodat, belépés után a saját eredményednél megteheted.
"""
            if found.email and found.email.lower() == contact:
                sent = send_code_email(found.email, "JudgeFlow – Versenyző kód emlékeztető", body)
                if sent:
                    flash("Email elküldve a regisztrált email címre!" if hu() else "Email sent!", "success")
                else:
                    flash("Email küldés sikertelen. Kérdezd meg a rendezőt!" if hu() else "Could not send email. Ask the organizer!", "error")
            else:
                # Phone - can't send SMS automatically, show organizer contact
                flash(f"Telefonra nem tudunk automatikusan kódot küldeni. Kérd a versenyrendezőt ({event.competition.organizer_name}, {event.competition.organizer_phone})!" if hu() else f"Cannot send SMS automatically. Contact the organizer: {event.competition.organizer_name}, {event.competition.organizer_phone}", "error")
        else:
            # Contact not found - show organizer info
            comp = event.competition
            flash(
                f"Nem találtuk az elérhetőséget. Kérdezd meg a versenyrendezőt: {comp.organizer_name} · {comp.organizer_phone} · {comp.organizer_email}" if hu() else
                f"Contact not found. Ask the organizer: {comp.organizer_name} · {comp.organizer_phone} · {comp.organizer_email}",
                "error"
            )
        return redirect(url_for('rider_event_login', event_id=event_id))
    return render_template('forgot_rider.html', event_id=event_id)

# ── Init ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=8080, debug=False)
